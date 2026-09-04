import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
import numpy as np
import pyomo.environ as pyo
from pyomo.opt import SolverFactory, TerminationCondition
from itertools import combinations
from datetime import datetime
import io

# ================================================================
# PAGE CONFIGURATION
# ================================================================
st.set_page_config(
    page_title="Phase 2 and 3: Criteria Selection Tool of CREST",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ================================================================
# CUSTOM CSS FOR PROFESSIONAL LOOK
# ================================================================
st.markdown("""
<style>
    .main-title { font-size: 2.5rem; font-weight: 700; color: #1f2937; margin-bottom: 0.5rem; }
    .sub-title { font-size: 1.25rem; color: #6b7280; margin-bottom: 2rem; }
    .reference-box { background: #fef3c7; border: 2px solid #f59e0b; padding: 1rem; border-radius: 8px; margin: 1rem 0; font-weight: 600; }
    .stButton button { border-radius: 8px; padding: 0.5rem 1.5rem; font-weight: 500; transition: all 0.3s; }
    div[data-testid="metric-container"] { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 1rem; border-radius: 10px; color: white; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
    div[data-testid="metric-container"] label { color: rgba(255,255,255,0.9) !important; font-size: 0.875rem !important; font-weight: 600 !important; }
    div[data-testid="metric-container"] [data-testid="stMetricValue"] { color: white !important; font-size: 2rem !important; font-weight: 700 !important; }
    section[data-testid="stSidebar"] { background: linear-gradient(180deg, #f8f9fa 0%, #e9ecef 100%); }
    .info-box { background: #eff6ff; border-left: 4px solid #3b82f6; padding: 1rem; border-radius: 8px; margin: 1rem 0; }
    .success-box { background: #f0fdf4; border-left: 4px solid #10b981; padding: 1rem; border-radius: 8px; margin: 1rem 0; }
    .warning-box { background: #fffbeb; border-left: 4px solid #f59e0b; padding: 1rem; border-radius: 8px; margin: 1rem 0; }
</style>
""", unsafe_allow_html=True)

# ================================================================
# SESSION STATE INITIALIZATION
# ================================================================
for _key, _default in [('data', None), ('weights', None), ('solution', None), ('result_frames', None),
                       ('config', None), ('current_step', 1), ('reference_component', 'w5_plus')]:
    if _key not in st.session_state:
        st.session_state[_key] = _default

# ================================================================
# PROPERTY DEFINITIONS (13 properties of the revised CREST framework)
# ================================================================
PROPERTIES = {
    1: "Completeness",
    2: "Alignment",
    3: "Directness",
    4: "Representativeness",
    5: "Parsimony",
    6: "Assessment Mode",
    7: "Operationality",
    8: "Understandability",
    9: "Cost-effectiveness",
    10: "Unambiguity",
    11: "Monotone Coherence",
    12: "Distinctiveness",
    13: "Sensitivity",
}

PROPERTY_EFFECTS = {
    1: "Two score gates and normalized reward",
    2: "Score gate and normalized reward",
    3: "Score gate and normalized reward",
    4: "Objective coverage and under/over-representation penalties",
    5: "Under/over-complexity penalties",
    6: "Quantitative-composition interval; rho is always reported",
    7: "Score gate and normalized reward",
    8: "Score gate and normalized reward",
    9: "Score gate and normalized reward",
    10: "Score gate and normalized reward",
    11: "Hard veto",
    12: "Pairwise exclusion and overlap penalty",
    13: "Score gate and normalized reward",
}

# Weight components of the objective function. Property 6 and 11 carry no weight.
WEIGHT_COMPONENTS = {
    'w1': (1, 'Completeness', 'Reward', 'Coverage of concerns and consequence ranges'),
    'w2': (2, 'Alignment', 'Reward', 'Relevance to the decision objectives'),
    'w3': (3, 'Directness', 'Reward', 'Directness of the link to the underlying objective'),
    'w4_minus': (4, 'Representativeness (under)', 'Penalty', 'Penalty for objectives represented below L(o)'),
    'w4_plus': (4, 'Representativeness (over)', 'Penalty', 'Penalty for objectives represented above U(o)'),
    'w5_minus': (5, 'Parsimony (under-complexity)', 'Penalty', 'Penalty for selecting fewer than omega criteria'),
    'w5_plus': (5, 'Parsimony (over-complexity)', 'Penalty', 'Penalty for selecting more than zeta criteria'),
    'w7': (7, 'Operationality', 'Reward', 'Consistency of practical assessment'),
    'w8': (8, 'Understandability', 'Reward', 'Clarity to the intended users'),
    'w9': (9, 'Cost-effectiveness', 'Reward', 'Information value relative to assessment burden'),
    'w10': (10, 'Unambiguity', 'Reward', 'Precision of consequence-to-level mapping'),
    'w12': (12, 'Distinctiveness', 'Penalty', 'Penalty for jointly selecting correlated criteria'),
    'w13': (13, 'Sensitivity', 'Reward', 'Influence of the criterion on the decision outcome'),
}
WEIGHT_NOTATION = {
    'w1': 'w1', 'w2': 'w2', 'w3': 'w3', 'w4_minus': 'w4^-', 'w4_plus': 'w4^+',
    'w5_minus': 'w5^-', 'w5_plus': 'w5^+', 'w7': 'w7', 'w8': 'w8', 'w9': 'w9',
    'w10': 'w10', 'w12': 'w12', 'w13': 'w13',
}

RATING_QUESTIONS = {
    2: "To what extent is criterion i relevant to the decision objectives, that is, does it capture a consequence that matters for the choice among alternatives? (0 to 10)",
    3: "To what extent does criterion i directly measure the consequence described by its designated underlying objective, rather than serving as an indirect proxy? (0 to 10)",
    7: "To what extent can criterion i be assessed consistently in practice using a clear assessment procedure and reasonably obtainable information? (0 to 10)",
    8: "To what extent can the intended users readily understand what criterion i means, how it is assessed, and what the resulting evaluations mean? (0 to 10)",
    9: "Judge the decision-relevant informational value of criterion i relative to the resources required for its assessment (financial cost, time, data, expertise, effort). Higher scores indicate a more favorable balance. (0 to 10)",
}

TOL = 1e-12

# ================================================================
# EXCEL TEMPLATE GENERATOR
# ================================================================

def generate_excel_template(num_criteria, num_alternatives, num_experts, num_objectives,
                            omega, zeta, L_list, U_list, active, thresholds, n_mc, seed, M_big, eps):
    """Generate the complete Excel template (Configuration plus one sheet per property).

    Inactive properties still receive a sheet, but their input cells are prefilled with 0 and the
    sheet is flagged INACTIVE. Their inputs are ignored by the app and their weights are locked to 0.
    Returns (buffer, config_dict).
    """

    config = {
        'num_criteria': num_criteria, 'num_alternatives': num_alternatives,
        'num_experts': num_experts, 'num_objectives': num_objectives,
        'omega': omega, 'zeta': zeta, 'L': L_list, 'U': U_list,
        'active': dict(active), 'thresholds': dict(thresholds),
        'n_mc': n_mc, 'seed': seed, 'M': M_big, 'eps': eps,
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    output_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    section_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    inactive_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    def header_row(ws, row, headers):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    def input_cell(ws, row, col, prop_active, value=None):
        cell = ws.cell(row=row, column=col)
        if prop_active:
            cell.fill = input_fill
            if value is not None:
                cell.value = value
        else:
            cell.fill = inactive_fill
            cell.value = 0
        cell.border = thin_border
        return cell

    def output_cell(ws, row, col, formula, fmt=None):
        cell = ws.cell(row=row, column=col, value=formula)
        cell.fill = output_fill
        cell.border = thin_border
        if fmt:
            cell.number_format = fmt
        return cell

    def sheet_banner(ws, title, prop_num, extra_line=""):
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=12)
        if active[prop_num]:
            ws['A2'] = extra_line if extra_line else "Status: ACTIVE"
        else:
            ws['A2'] = "Status: INACTIVE. This property is not applied. Cells are prefilled with 0 and ignored by the app; its weight is locked to 0."
            ws['A2'].font = Font(bold=True, color="C00000")
        ws.merge_cells('A2:J2')

    # ------------------------------------------------------------
    # SHEET: CONFIGURATION
    # ------------------------------------------------------------
    ws_config = wb.create_sheet("Configuration")
    ws_config['A1'] = "CREST CRITERIA SELECTION - CONFIGURATION"
    ws_config['A1'].font = Font(bold=True, size=14)
    ws_config.merge_cells('A1:E1')

    meta = {}
    row = 3
    ws_config[f'A{row}'] = "PROBLEM STRUCTURE"
    ws_config[f'A{row}'].font = Font(bold=True, size=12)
    ws_config[f'A{row}'].fill = section_fill
    ws_config.merge_cells(f'A{row}:E{row}')
    row += 1
    for label, value in [["Number of Criteria", num_criteria], ["Number of Alternatives", num_alternatives],
                         ["Number of Experts", num_experts], ["Number of Objectives", num_objectives]]:
        ws_config[f'A{row}'] = label
        ws_config[f'B{row}'] = value
        row += 1
    row += 1

    ws_config[f'A{row}'] = "CRITERIA DEFINITIONS (Fill in the yellow cells)"
    ws_config[f'A{row}'].font = Font(bold=True, size=12)
    ws_config[f'A{row}'].fill = section_fill
    ws_config.merge_cells(f'A{row}:E{row}')
    row += 1
    header_row(ws_config, row, ["Criterion ID", "Criterion Name", "Type (Cost/Benefit)", "Description (Optional)"])
    row += 1
    CRITERIA_START_ROW = row
    meta['criteria_start_row'] = CRITERIA_START_ROW
    for i in range(num_criteria):
        ws_config.cell(row=row, column=1, value=f"C{i+1}")
        c = ws_config.cell(row=row, column=2, value=f"Criterion {i+1}"); c.fill = input_fill; c.border = thin_border
        c = ws_config.cell(row=row, column=3, value="Benefit"); c.fill = input_fill; c.border = thin_border
        c = ws_config.cell(row=row, column=4, value=""); c.fill = input_fill; c.border = thin_border
        row += 1
    dv = DataValidation(type="list", formula1='"Cost,Benefit"', allow_blank=False)
    ws_config.add_data_validation(dv)
    dv.add(f"C{CRITERIA_START_ROW}:C{CRITERIA_START_ROW + num_criteria - 1}")
    row += 1

    ws_config[f'A{row}'] = "ALTERNATIVES DEFINITIONS (Fill in the yellow cells)"
    ws_config[f'A{row}'].font = Font(bold=True, size=12)
    ws_config[f'A{row}'].fill = section_fill
    ws_config.merge_cells(f'A{row}:E{row}')
    row += 1
    header_row(ws_config, row, ["Alternative ID", "Alternative Name", "Description (Optional)"])
    row += 1
    ALTERNATIVES_START_ROW = row
    meta['alternatives_start_row'] = ALTERNATIVES_START_ROW
    for i in range(num_alternatives):
        ws_config.cell(row=row, column=1, value=f"A{i+1}")
        c = ws_config.cell(row=row, column=2, value=f"Alternative {i+1}"); c.fill = input_fill; c.border = thin_border
        c = ws_config.cell(row=row, column=3, value=""); c.fill = input_fill; c.border = thin_border
        row += 1
    row += 1

    ws_config[f'A{row}'] = "OBJECTIVES DEFINITIONS AND REPRESENTATION TARGETS (Property IV)"
    ws_config[f'A{row}'].font = Font(bold=True, size=12)
    ws_config[f'A{row}'].fill = section_fill
    ws_config.merge_cells(f'A{row}:E{row}')
    row += 1
    header_row(ws_config, row, ["Objective ID", "Objective Name", "Min representation L(o)", "Max representation U(o)", "Description (Optional)"])
    row += 1
    OBJECTIVES_START_ROW = row
    meta['objectives_start_row'] = OBJECTIVES_START_ROW
    for o in range(num_objectives):
        ws_config.cell(row=row, column=1, value=f"O{o+1}")
        c = ws_config.cell(row=row, column=2, value=f"Objective {o+1}"); c.fill = input_fill; c.border = thin_border
        c = ws_config.cell(row=row, column=3, value=int(L_list[o])); c.fill = input_fill; c.border = thin_border
        c = ws_config.cell(row=row, column=4, value=int(U_list[o])); c.fill = input_fill; c.border = thin_border
        c = ws_config.cell(row=row, column=5, value=""); c.fill = input_fill; c.border = thin_border
        row += 1
    row += 1

    ws_config[f'A{row}'] = "PROPERTY ACTIVATION (set in the app; Yes = applied, No = ignored and weight locked to 0)"
    ws_config[f'A{row}'].font = Font(bold=True, size=12)
    ws_config[f'A{row}'].fill = section_fill
    ws_config.merge_cells(f'A{row}:E{row}')
    row += 1
    header_row(ws_config, row, ["Property no.", "Property", "Use property?", "Effect when active"])
    row += 1
    ACTIVATION_START_ROW = row
    meta['activation_start_row'] = ACTIVATION_START_ROW
    for p in range(1, 14):
        ws_config.cell(row=row, column=1, value=p)
        ws_config.cell(row=row, column=2, value=PROPERTIES[p])
        c = ws_config.cell(row=row, column=3, value="Yes" if active[p] else "No")
        c.border = thin_border
        c.fill = output_fill if active[p] else inactive_fill
        ws_config.cell(row=row, column=4, value=PROPERTY_EFFECTS[p])
        row += 1
    row += 1

    ws_config[f'A{row}'] = "PARSIMONY TARGETS (Property V)"
    ws_config[f'A{row}'].font = Font(bold=True, size=12)
    ws_config[f'A{row}'].fill = section_fill
    ws_config.merge_cells(f'A{row}:E{row}')
    row += 1
    PARSIMONY_ROW = row
    meta['parsimony_row'] = PARSIMONY_ROW
    ws_config[f'A{row}'] = "Target Minimum (omega)"; ws_config[f'B{row}'] = omega; row += 1
    ws_config[f'A{row}'] = "Target Maximum (zeta)"; ws_config[f'B{row}'] = zeta; row += 1
    row += 1

    ws_config[f'A{row}'] = "THRESHOLDS, BOUNDS, AND MODEL SETTINGS"
    ws_config[f'A{row}'].font = Font(bold=True, size=12)
    ws_config[f'A{row}'].fill = section_fill
    ws_config.merge_cells(f'A{row}:E{row}')
    row += 1
    THRESHOLD_START_ROW = row
    meta['threshold_start_row'] = THRESHOLD_START_ROW
    threshold_rows = [
        ("alpha_con", "Property I: Completeness, concern coverage (alpha^con)"),
        ("alpha_rng", "Property I: Completeness, consequence range (alpha^rng)"),
        ("lambda", "Property II: Alignment (lambda)"),
        ("psi", "Property III: Directness (psi)"),
        ("rho_LB", "Property VI: Assessment Mode lower bound (rho^LB)"),
        ("rho_UB", "Property VI: Assessment Mode upper bound (rho^UB)"),
        ("gamma", "Property VII: Operationality (gamma)"),
        ("eta", "Property VIII: Understandability (eta)"),
        ("tau", "Property IX: Cost-effectiveness (tau)"),
        ("mu", "Property X: Unambiguity (mu)"),
        ("delta", "Property XII: Distinctiveness (delta)"),
        ("theta", "Property XIII: Sensitivity (theta)"),
        ("n_mc", "Property XIII: Monte Carlo runs (N_MC)"),
        ("seed", "Property XIII: Random seed"),
        ("M", "MILP Big-M constant (M)"),
        ("eps", "MILP strict-comparison tolerance (epsilon)"),
    ]
    values_map = dict(thresholds)
    values_map.update({'n_mc': n_mc, 'seed': seed, 'M': M_big, 'eps': eps})
    meta['threshold_keys'] = ",".join(k for k, _ in threshold_rows)
    for key, label in threshold_rows:
        ws_config[f'A{row}'] = label
        ws_config[f'B{row}'] = values_map[key]
        ws_config[f'C{row}'] = key
        row += 1

    ws_config.column_dimensions['A'].width = 55
    ws_config.column_dimensions['B'].width = 30
    ws_config.column_dimensions['C'].width = 22
    ws_config.column_dimensions['D'].width = 22
    ws_config.column_dimensions['E'].width = 30

    def name_ref(i):
        return f'=Configuration!$B${CRITERIA_START_ROW + i}'

    def type_ref(i):
        return f'=Configuration!$C${CRITERIA_START_ROW + i}'

    def alt_ref(a):
        return f'=Configuration!$B${ALTERNATIVES_START_ROW + a}'

    E = num_experts
    DATA_START = 6

    # ------------------------------------------------------------
    # SHEET I: COMPLETENESS (two dimensions)
    # ------------------------------------------------------------
    ws1 = wb.create_sheet("I_Completeness")
    sheet_banner(ws1, "Property I: Completeness Evaluation (Concern coverage and Consequence-range coverage)", 1,
                 f"Thresholds: alpha^con = {thresholds['alpha_con']}, alpha^rng = {thresholds['alpha_rng']}. Overall Completeness = min(median concern, median range); both gates must be met.")
    ws1['A3'] = ("Concern coverage: considering the decision context and the current set, if criterion i were excluded, how much important decision-relevant concern would be missed? (0 to 10). "
                 "Consequence-range coverage: to what extent does the assessment scale of criterion i span the full range of consequences that can realistically occur across alternatives? (0 to 10)")
    ws1.merge_cells('A3:L3')
    ws1['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    headers = ["Criterion ID", "Criterion Name"] + [f"Concern E{e+1}" for e in range(E)] + [f"Range E{e+1}" for e in range(E)]
    headers += ["Median Concern", "Median Range", "Overall (min)", "Concern Status", "Range Status"]
    header_row(ws1, 5, headers)
    for i in range(num_criteria):
        r = DATA_START + i
        ws1.cell(row=r, column=1, value=f"C{i+1}")
        c = ws1.cell(row=r, column=2, value=name_ref(i)); c.border = thin_border
        for e in range(E):
            input_cell(ws1, r, 3 + e, active[1])
            input_cell(ws1, r, 3 + E + e, active[1])
        c1, c2 = get_column_letter(3), get_column_letter(2 + E)
        r1, r2 = get_column_letter(3 + E), get_column_letter(2 + 2 * E)
        mc_col, mr_col = 3 + 2 * E, 4 + 2 * E
        output_cell(ws1, r, mc_col, f'=MEDIAN({c1}{r}:{c2}{r})', '0.00')
        output_cell(ws1, r, mr_col, f'=MEDIAN({r1}{r}:{r2}{r})', '0.00')
        mcl, mrl = get_column_letter(mc_col), get_column_letter(mr_col)
        output_cell(ws1, r, 5 + 2 * E, f'=MIN({mcl}{r},{mrl}{r})', '0.00')
        output_cell(ws1, r, 6 + 2 * E, f'=IF({mcl}{r}>={thresholds["alpha_con"]},"Meets","Below")')
        output_cell(ws1, r, 7 + 2 * E, f'=IF({mrl}{r}>={thresholds["alpha_rng"]},"Meets","Below")')
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 30
    for j in range(2 * E + 5):
        ws1.column_dimensions[get_column_letter(3 + j)].width = 13
    ws1.row_dimensions[3].height = 75

    # ------------------------------------------------------------
    # GENERIC RATING SHEETS: II, III, VII, VIII, IX
    # ------------------------------------------------------------
    rating_specs = [
        (2, "II_Alignment", "Property II: Alignment Assessment", "lambda", "lambda"),
        (3, "III_Directness", "Property III: Directness Assessment", "psi", "psi"),
        (7, "VII_Operationality", "Property VII: Operationality Assessment", "gamma", "gamma"),
        (8, "VIII_Understandability", "Property VIII: Understandability Assessment", "eta", "eta"),
        (9, "IX_Cost_Effectiveness", "Property IX: Cost-Effectiveness Evaluation", "tau", "tau"),
    ]
    for prop, sheet_name, title, thr_key, thr_symbol in rating_specs:
        ws = wb.create_sheet(sheet_name)
        thr = thresholds[thr_key]
        sheet_banner(ws, title, prop, f"Threshold: {thr_symbol} = {thr}")
        ws['A3'] = RATING_QUESTIONS[prop]
        ws.merge_cells('A3:J3')
        ws['A3'].alignment = Alignment(wrap_text=True, vertical='top')
        extra = ["Designated Objective (optional)"] if prop == 3 else []
        offset = 3 + len(extra)
        headers = ["Criterion ID", "Criterion Name"] + extra + [f"Expert {e+1}" for e in range(E)] + ["Median", "Status"]
        header_row(ws, 5, headers)
        for i in range(num_criteria):
            r = DATA_START + i
            ws.cell(row=r, column=1, value=f"C{i+1}")
            c = ws.cell(row=r, column=2, value=name_ref(i)); c.border = thin_border
            if prop == 3:
                c = ws.cell(row=r, column=3, value=""); c.fill = input_fill; c.border = thin_border
            for e in range(E):
                input_cell(ws, r, offset + e, active[prop])
            first, last = get_column_letter(offset), get_column_letter(offset + E - 1)
            med_col = offset + E
            output_cell(ws, r, med_col, f'=MEDIAN({first}{r}:{last}{r})', '0.00')
            output_cell(ws, r, med_col + 1, f'=IF({get_column_letter(med_col)}{r}>={thr},"Meets","Below")')
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        if prop == 3:
            ws.column_dimensions['C'].width = 26
        for j in range(E + 2):
            ws.column_dimensions[get_column_letter(offset + j)].width = 12
        ws.row_dimensions[3].height = 60

    # ------------------------------------------------------------
    # SHEET IV: REPRESENTATIVENESS
    # ------------------------------------------------------------
    ws4 = wb.create_sheet("IV_Representativeness")
    sheet_banner(ws4, "Property IV: Representativeness (criterion to objective assignments)", 4,
                 "L(o) and U(o) per objective are set in the Configuration sheet. Consolidation uses strict majority: g_io = 1 only when more than half of the experts vote 1.")
    ws4['A3'] = "For each criterion i and each finalized objective o, indicate whether criterion i meaningfully represents objective o (1 = yes, 0 = no). Do not leave cells blank. A criterion may represent several objectives."
    ws4.merge_cells('A3:J3')
    ws4['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    rep_expert_rows = []
    row = 5
    for e in range(E):
        ws4.cell(row=row, column=1, value=f"Expert {e+1} Assignments").font = Font(bold=True)
        row += 1
        header_row(ws4, row, ["Criterion"] + [f"O{o+1}" for o in range(num_objectives)])
        row += 1
        rep_expert_rows.append(row)
        for i in range(num_criteria):
            c = ws4.cell(row=row, column=1, value=name_ref(i)); c.border = thin_border
            for o in range(num_objectives):
                input_cell(ws4, row, 2 + o, active[4])
            row += 1
        row += 2
    meta['rep_expert_rows'] = ",".join(str(r) for r in rep_expert_rows)
    row += 1
    ws4.cell(row=row, column=1, value="CONSOLIDATED (strict majority)").font = Font(bold=True, size=12)
    row += 2
    header_row(ws4, row, ["Criterion"] + [f"O{o+1}" for o in range(num_objectives)])
    for i in range(num_criteria):
        row += 1
        c = ws4.cell(row=row, column=1, value=name_ref(i)); c.border = thin_border
        for o in range(num_objectives):
            col = get_column_letter(2 + o)
            refs = "+".join(f"{col}{rep_expert_rows[e] + i}" for e in range(E))
            output_cell(ws4, row, 2 + o, f'=IF({refs}>{E}/2,1,0)')
    row += 2
    header_row(ws4, row, ["Objective", "|I_o| (candidates)"])
    cons_first = row - num_criteria - 1
    for o in range(num_objectives):
        row += 1
        ws4.cell(row=row, column=1, value=f"O{o+1}")
        col = get_column_letter(2 + o)
        output_cell(ws4, row, 2, f'=SUM({col}{cons_first + 1}:{col}{cons_first + num_criteria})')
    ws4.column_dimensions['A'].width = 35
    for o in range(num_objectives + 1):
        ws4.column_dimensions[get_column_letter(2 + o)].width = 10
    ws4.row_dimensions[3].height = 50

    # ------------------------------------------------------------
    # SHEET VI: ASSESSMENT MODE
    # ------------------------------------------------------------
    ws6 = wb.create_sheet("VI_Assessment_Mode")
    sheet_banner(ws6, "Property VI: Assessment Mode Classification", 6,
                 f"Composition interval when active: rho^LB = {thresholds['rho_LB']}, rho^UB = {thresholds['rho_UB']}. The quantitative proportion rho of the selected family is always reported.")
    ws6['A3'] = "Is criterion i assessed primarily using measured or calculated quantitative values (1), or primarily through qualitative or ordinal judgment (0)? Strict majority is required; an even-panel tie must be re-elicited."
    ws6.merge_cells('A3:J3')
    ws6['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    header_row(ws6, 5, ["Criterion ID", "Criterion Name"] + [f"Expert {e+1}" for e in range(E)] + ["Votes (quantitative)", "Tag o_i"])
    for i in range(num_criteria):
        r = DATA_START + i
        ws6.cell(row=r, column=1, value=f"C{i+1}")
        c = ws6.cell(row=r, column=2, value=name_ref(i)); c.border = thin_border
        for e in range(E):
            input_cell(ws6, r, 3 + e, active[6])
        first, last = get_column_letter(3), get_column_letter(2 + E)
        output_cell(ws6, r, 3 + E, f'=SUM({first}{r}:{last}{r})')
        s = get_column_letter(3 + E)
        output_cell(ws6, r, 4 + E, f'=IF({s}{r}>{E}/2,1,IF({s}{r}<{E}/2,0,"TIE"))')
    ws6.column_dimensions['A'].width = 12
    ws6.column_dimensions['B'].width = 30
    for j in range(E + 2):
        ws6.column_dimensions[get_column_letter(3 + j)].width = 14
    ws6.row_dimensions[3].height = 50

    # ------------------------------------------------------------
    # SHEET X: UNAMBIGUITY (cross-expert ratings)
    # ------------------------------------------------------------
    ws10 = wb.create_sheet("X_Unambiguity")
    sheet_banner(ws10, "Property X: Unambiguity (cross-expert ratings of mapping explanations)", 10,
                 f"Threshold: mu = {thresholds['mu']}. Self-ratings are excluded; the median is taken over all cross-expert ratings.")
    ws10['A3'] = ("Each expert authored an explanation of how consequences are mapped onto the assessment levels of criterion i. Every other expert rates that explanation: "
                  "to what extent does it specify a precise and assessor-independent mapping from consequences to levels? (0 to 10). Column E_r to E_a = rating by expert r of the explanation authored by expert a.")
    ws10.merge_cells('A3:L3')
    ws10['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    cross_headers = [f"E{r+1} to E{a+1}" for r in range(E) for a in range(E) if r != a]
    n_cross = len(cross_headers)
    header_row(ws10, 5, ["Criterion ID", "Criterion Name"] + cross_headers + ["Median", "Status"])
    for i in range(num_criteria):
        r = DATA_START + i
        ws10.cell(row=r, column=1, value=f"C{i+1}")
        c = ws10.cell(row=r, column=2, value=name_ref(i)); c.border = thin_border
        for j in range(n_cross):
            input_cell(ws10, r, 3 + j, active[10])
        if n_cross > 0:
            first, last = get_column_letter(3), get_column_letter(2 + n_cross)
            output_cell(ws10, r, 3 + n_cross, f'=MEDIAN({first}{r}:{last}{r})', '0.00')
            output_cell(ws10, r, 4 + n_cross, f'=IF({get_column_letter(3 + n_cross)}{r}>={thresholds["mu"]},"Meets","Below")')
    ws10.column_dimensions['A'].width = 12
    ws10.column_dimensions['B'].width = 30
    for j in range(n_cross + 2):
        ws10.column_dimensions[get_column_letter(3 + j)].width = 11
    ws10.row_dimensions[3].height = 75

    # ------------------------------------------------------------
    # SHEET XI: MONOTONE COHERENCE
    # ------------------------------------------------------------
    ws11 = wb.create_sheet("XI_Monotone_Coherence")
    sheet_banner(ws11, "Property XI: Monotone Coherence", 11,
                 "q_i = product of expert responses (unanimity). When active, a criterion with q_i = 0 is vetoed.")
    ws11['A3'] = "Holding all other criteria fixed, does movement in the stated preferred direction on criterion i (an increase for a benefit criterion, a decrease for a cost criterion) never make an alternative less preferred? (1 = yes, 0 = no)"
    ws11.merge_cells('A3:J3')
    ws11['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    header_row(ws11, 5, ["Criterion ID", "Criterion Name", "Type", "Preferred direction"] + [f"Expert {e+1}" for e in range(E)] + ["q_i", "Status"])
    for i in range(num_criteria):
        r = DATA_START + i
        ws11.cell(row=r, column=1, value=f"C{i+1}")
        c = ws11.cell(row=r, column=2, value=name_ref(i)); c.border = thin_border
        c = ws11.cell(row=r, column=3, value=type_ref(i)); c.border = thin_border
        c = ws11.cell(row=r, column=4, value=f'=IF(C{r}="Cost","decrease","increase")'); c.border = thin_border
        for e in range(E):
            input_cell(ws11, r, 5 + e, active[11])
        first, last = get_column_letter(5), get_column_letter(4 + E)
        output_cell(ws11, r, 5 + E, f'=PRODUCT({first}{r}:{last}{r})')
        output_cell(ws11, r, 6 + E, f'=IF({get_column_letter(5 + E)}{r}=1,"Meets","Vetoed")')
    ws11.column_dimensions['A'].width = 12
    ws11.column_dimensions['B'].width = 30
    ws11.column_dimensions['C'].width = 10
    ws11.column_dimensions['D'].width = 18
    for j in range(E + 2):
        ws11.column_dimensions[get_column_letter(5 + j)].width = 11
    ws11.row_dimensions[3].height = 60

    # ------------------------------------------------------------
    # SHEETS XII and XIII: DECISION MATRICES
    # ------------------------------------------------------------
    def matrix_sheet(sheet_name, title, prop_num, line2, line3):
        ws = wb.create_sheet(sheet_name)
        sheet_banner(ws, title, prop_num, line2)
        ws['A3'] = line3
        ws.merge_cells('A3:J3')
        ws['A3'].alignment = Alignment(wrap_text=True, vertical='top')
        row = 6
        block_rows = []
        for e in range(E):
            ws.cell(row=row, column=1, value=f"Expert {e+1} Decision Matrix").font = Font(bold=True)
            row += 1
            header_row(ws, row, ["Alternative"] + [f"C{c+1}" for c in range(num_criteria)])
            row += 1
            block_rows.append(row)
            for a in range(num_alternatives):
                c = ws.cell(row=row, column=1, value=alt_ref(a)); c.border = thin_border
                for k in range(num_criteria):
                    input_cell(ws, row, 2 + k, active[prop_num])
                row += 1
            row += 2
        ws.column_dimensions['A'].width = 35
        for k in range(num_criteria):
            ws.column_dimensions[get_column_letter(2 + k)].width = 10
        ws.row_dimensions[3].height = 60
        return block_rows

    rows12 = matrix_sheet("XII_Distinctiveness", "Property XII: Distinctiveness - Decision Matrices", 12,
                          f"Correlation threshold: delta = {thresholds['delta']}",
                          "Each expert provides an alternative-by-criterion performance matrix (raw performances). The app computes the absolute Pearson correlation of every criterion pair within each expert matrix and pools them by the median.")
    rows13 = matrix_sheet("XIII_Sensitivity", "Property XIII: Sensitivity - Decision Matrices", 13,
                          f"Sensitivity threshold: theta = {thresholds['theta']}; N_MC = {n_mc}; seed = {seed}",
                          "Each expert provides an alternative-by-criterion performance matrix. The app applies direction-aware normalization, draws N_MC Dirichlet weight vectors, and computes the average relative influence of each criterion.")
    meta['matrix_rows_12'] = ",".join(str(r) for r in rows12)
    meta['matrix_rows_13'] = ",".join(str(r) for r in rows13)

    # ------------------------------------------------------------
    # HIDDEN META SHEET (row positions used by the reader)
    # ------------------------------------------------------------
    ws_meta = wb.create_sheet("_Meta")
    ws_meta['A1'] = "Key"; ws_meta['B1'] = "Value"
    meta.update({
        'template_version': "CREST-13",
        'num_criteria': num_criteria, 'num_alternatives': num_alternatives,
        'num_experts': num_experts, 'num_objectives': num_objectives,
    })
    for r, (k, v) in enumerate(meta.items(), start=2):
        ws_meta.cell(row=r, column=1, value=k)
        ws_meta.cell(row=r, column=2, value=v)
    ws_meta.sheet_state = 'hidden'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, config

# ================================================================
# EXCEL READER AND PROPERTY COMPUTATIONS
# ================================================================

def _num(v, where):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        raise ValueError(f"Missing numeric value at {where}.")
    try:
        return float(v)
    except (TypeError, ValueError):
        raise ValueError(f"Non-numeric value {v!r} at {where}.")


def read_mcdm_template(file):
    """Read the filled template and compute all model-ready property values.

    Replicates the manuscript procedures: medians for score properties, min for Completeness,
    strict majority for Representativeness and Assessment Mode, unanimity product for Monotone
    Coherence, pooled median absolute Pearson correlation for Distinctiveness, and the Dirichlet
    Monte Carlo relative-influence procedure for Sensitivity.
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    if "_Meta" not in wb.sheetnames:
        raise ValueError("This file is not a CREST template generated by this app (missing _Meta sheet).")
    meta = {}
    for r in wb["_Meta"].iter_rows(min_row=2, values_only=True):
        if r[0] is not None:
            meta[str(r[0])] = r[1]
    if meta.get('template_version') != "CREST-13":
        raise ValueError("Template version mismatch. Please regenerate the template with this version of the app.")

    n = int(meta['num_criteria']); A = int(meta['num_alternatives'])
    E = int(meta['num_experts']); nO = int(meta['num_objectives'])
    cfg = wb["Configuration"]

    crit_row = int(meta['criteria_start_row'])
    criteria_names, criteria_types = [], []
    for i in range(n):
        name = cfg.cell(row=crit_row + i, column=2).value
        ctype = str(cfg.cell(row=crit_row + i, column=3).value or "Benefit").strip().capitalize()
        if ctype not in ("Cost", "Benefit"):
            raise ValueError(f"Criterion C{i+1}: type must be Cost or Benefit, found {ctype!r}.")
        criteria_names.append(str(name) if name is not None else f"Criterion {i+1}")
        criteria_types.append(ctype)

    alt_row = int(meta['alternatives_start_row'])
    alternatives = [str(cfg.cell(row=alt_row + a, column=2).value or f"Alternative {a+1}") for a in range(A)]

    obj_row = int(meta['objectives_start_row'])
    objectives_names, objectives_desc, L, U = [], [], {}, {}
    for o in range(nO):
        objectives_names.append(str(cfg.cell(row=obj_row + o, column=2).value or f"Objective {o+1}"))
        L[o + 1] = int(_num(cfg.cell(row=obj_row + o, column=3).value, f"L(O{o+1})"))
        U[o + 1] = int(_num(cfg.cell(row=obj_row + o, column=4).value, f"U(O{o+1})"))
        objectives_desc.append(str(cfg.cell(row=obj_row + o, column=5).value or ""))

    act_row = int(meta['activation_start_row'])
    active = {}
    for p in range(1, 14):
        v = str(cfg.cell(row=act_row + p - 1, column=3).value or "").strip().lower()
        if v not in ("yes", "no"):
            raise ValueError(f"Property {p} activation must be Yes or No in the Configuration sheet.")
        active[p] = (v == "yes")

    par_row = int(meta['parsimony_row'])
    omega = int(_num(cfg.cell(row=par_row, column=2).value, "omega"))
    zeta = int(_num(cfg.cell(row=par_row + 1, column=2).value, "zeta"))

    thr_row = int(meta['threshold_start_row'])
    thr_keys = str(meta['threshold_keys']).split(",")
    thr = {}
    for j, key in enumerate(thr_keys):
        thr[key] = _num(cfg.cell(row=thr_row + j, column=2).value, key)
    n_mc, seed = int(thr.pop('n_mc')), int(thr.pop('seed'))
    M_big, eps = thr.pop('M'), thr.pop('eps')

    I = list(range(1, n + 1))
    O = list(range(1, nO + 1))
    DATA_START = 6

    def read_block(ws, col_offset, width, prop, lenient=False):
        """Return list (per criterion) of lists (width values) read from a rating sheet.

        Inactive properties are read as 0. With lenient=True (Assessment Mode), values are read
        even when the property is inactive so that rho can still be reported; blanks count as 0.
        """
        out = []
        for i in range(n):
            vals = []
            for j in range(width):
                cell = ws.cell(row=DATA_START + i, column=col_offset + j)
                if active[prop]:
                    vals.append(_num(cell.value, f"{ws.title}!{cell.coordinate}"))
                elif lenient:
                    try:
                        vals.append(_num(cell.value, ""))
                    except ValueError:
                        vals.append(0.0)
                else:
                    vals.append(0.0)
            out.append(vals)
        return out

    def median_scores(ws, col_offset, prop):
        block = read_block(ws, col_offset, E, prop)
        return {i: float(np.median(block[i - 1])) for i in I}, block

    # Property I
    ws1 = wb["I_Completeness"]
    concern = read_block(ws1, 3, E, 1)
    rng = read_block(ws1, 3 + E, E, 1)
    c_con = {i: float(np.median(concern[i - 1])) for i in I}
    c_rng = {i: float(np.median(rng[i - 1])) for i in I}
    c = {i: min(c_con[i], c_rng[i]) for i in I}

    # Properties II, III, VII, VIII, IX
    a, a_raw = median_scores(wb["II_Alignment"], 3, 2)
    dr, dr_raw = median_scores(wb["III_Directness"], 4, 3)
    designated_obj = [str(wb["III_Directness"].cell(row=DATA_START + i, column=3).value or "") for i in range(n)]
    op, op_raw = median_scores(wb["VII_Operationality"], 3, 7)
    un, un_raw = median_scores(wb["VIII_Understandability"], 3, 8)
    ce, ce_raw = median_scores(wb["IX_Cost_Effectiveness"], 3, 9)

    # Property IV: Representativeness (strict majority)
    ws4 = wb["IV_Representativeness"]
    rep_rows = [int(x) for x in str(meta['rep_expert_rows']).split(",")]
    votes = {(i, o): 0 for i in I for o in O}
    if active[4]:
        for e in range(E):
            for i in I:
                for o in O:
                    v = _num(ws4.cell(row=rep_rows[e] + i - 1, column=1 + o).value, f"IV_Representativeness expert {e+1} C{i} O{o}")
                    if v not in (0.0, 1.0):
                        raise ValueError(f"Representativeness votes must be 0 or 1 (expert {e+1}, C{i}, O{o}).")
                    votes[(i, o)] += int(v)
    g = {(i, o): (1 if votes[(i, o)] > E / 2 else 0) for i in I for o in O}
    I_o = {o: sum(g[(i, o)] for i in I) for o in O}
    D = {o: max(1, I_o[o] - U[o]) for o in O}
    if active[4]:
        for o in O:
            if I_o[o] < 1:
                raise ValueError(f"Objective O{o} ({objectives_names[o-1]}) has no candidate representative after strict-majority consolidation.")
            if not (1 <= L[o] <= U[o] <= I_o[o]):
                raise ValueError(f"Objective O{o}: representation targets must satisfy 1 <= L(o) <= U(o) <= |I_o|. Found L={L[o]}, U={U[o]}, |I_o|={I_o[o]}.")
    if active[5] and not (0 <= omega < zeta <= n):
        raise ValueError(f"Parsimony targets must satisfy 0 <= omega < zeta <= |I|. Found omega={omega}, zeta={zeta}, |I|={n}.")

    # Property VI: Assessment Mode (strict majority; ties unresolved)
    ws6 = wb["VI_Assessment_Mode"]
    o_votes = read_block(ws6, 3, E, 6, lenient=True)
    o_i, ties = {}, []
    for i in I:
        s = sum(o_votes[i - 1])
        if any(v not in (0.0, 1.0) for v in o_votes[i - 1]) and active[6]:
            raise ValueError(f"Assessment Mode votes must be 0 or 1 (C{i}).")
        if s > E / 2:
            o_i[i] = 1
        elif s < E / 2:
            o_i[i] = 0
        else:
            o_i[i] = 0
            ties.append(f"C{i} ({criteria_names[i-1]})")
    if ties and active[6]:
        raise ValueError("Assessment Mode is active but the strict-majority vote is tied for: " + "; ".join(ties) +
                         ". Re-elicit these criteria (or use an odd number of experts) and upload again.")
    rho_LB, rho_UB = thr['rho_LB'], thr['rho_UB']
    if active[6] and not (0 <= rho_LB <= rho_UB <= 1):
        raise ValueError("Assessment Mode bounds must satisfy 0 <= rho^LB <= rho^UB <= 1.")

    # Property X: Unambiguity (cross-expert median)
    ws10 = wb["X_Unambiguity"]
    n_cross = E * (E - 1)
    if active[10] and E < 2:
        raise ValueError("Unambiguity requires at least two experts. Deactivate Property X or add experts.")
    ua_raw = read_block(ws10, 3, n_cross, 10) if n_cross > 0 else [[0.0] for _ in I]
    ua = {i: float(np.median(ua_raw[i - 1])) for i in I}

    # Property XI: Monotone Coherence (unanimity)
    ws11 = wb["XI_Monotone_Coherence"]
    b_raw = read_block(ws11, 5, E, 11)
    q = {}
    for i in I:
        if active[11] and any(v not in (0.0, 1.0) for v in b_raw[i - 1]):
            raise ValueError(f"Monotone Coherence responses must be 0 or 1 (C{i}).")
        q[i] = int(all(v == 1.0 for v in b_raw[i - 1])) if active[11] else 1

    # Decision matrices for XII and XIII
    def read_matrices(ws, rows_key, prop):
        rows = [int(x) for x in str(meta[rows_key]).split(",")]
        mats = []
        for e in range(E):
            m = np.zeros((A, n))
            for aa in range(A):
                for k in range(n):
                    cell = ws.cell(row=rows[e] + aa, column=2 + k)
                    m[aa, k] = _num(cell.value, f"{ws.title}!{cell.coordinate}") if active[prop] else 0.0
            mats.append(m)
        return mats

    mats12 = read_matrices(wb["XII_Distinctiveness"], 'matrix_rows_12', 12)
    mats13 = read_matrices(wb["XIII_Sensitivity"], 'matrix_rows_13', 13)

    # Property XII: pooled median absolute Pearson correlation
    pairs = {(i, k): 0.0 for i, k in combinations(I, 2)}
    r_expert = {}
    if active[12]:
        if A < 2:
            raise ValueError("Distinctiveness requires at least two alternatives.")
        for e, m in enumerate(mats12):
            with np.errstate(invalid='ignore', divide='ignore'):
                corr = np.corrcoef(m, rowvar=False)
            if np.isnan(corr).any():
                bad = [f"C{k+1}" for k in range(n) if np.isclose(m[:, k].std(), 0.0)]
                raise ValueError(f"Expert {e+1} Distinctiveness matrix yields an undefined correlation. Constant criteria: {bad}.")
            r_expert[e] = np.abs(corr)
        for (i, k) in pairs:
            pairs[(i, k)] = float(np.median([r_expert[e][i - 1, k - 1] for e in range(E)]))

    # Property XIII: Monte Carlo relative influence
    s_bar = {i: 0.0 for i in I}
    if active[13]:
        if A < 2:
            raise ValueError("Sensitivity requires at least two alternatives.")
        rs = np.random.RandomState(seed)
        weights_mc = rs.dirichlet(np.ones(n), size=n_mc)
        expert_avgs = []
        for e, m in enumerate(mats13):
            norm = np.zeros_like(m)
            for k in range(n):
                vals = m[:, k]
                mn, mx = float(vals.min()), float(vals.max())
                alpha_, beta_ = (mx, mn) if criteria_types[k] == "Benefit" else (mn, mx)
                if np.isclose(alpha_, beta_):
                    raise ValueError(f"Sensitivity normalization requires variation across alternatives, but C{k+1} is constant for expert {e+1}.")
                norm[:, k] = np.abs(vals - beta_) / abs(alpha_ - beta_)
            alt_scores = norm @ weights_mc.T
            denom = alt_scores.sum(axis=0)
            if np.any(denom <= 0):
                raise ValueError(f"Total weighted alternative score is non-positive for expert {e+1} in at least one Monte Carlo run.")
            impacts = weights_mc * norm.sum(axis=0)[None, :]
            sens = impacts / denom[:, None]
            expert_avgs.append(sens.mean(axis=0))
        bar = np.mean(np.vstack(expert_avgs), axis=0)
        s_bar = {i: float(bar[i - 1]) for i in I}

    thresholds = {
        'alpha_con': thr['alpha_con'], 'alpha_rng': thr['alpha_rng'], 'lambda': thr['lambda'], 'psi': thr['psi'],
        'gamma': thr['gamma'], 'eta': thr['eta'], 'tau': thr['tau'], 'mu': thr['mu'],
        'delta': thr['delta'], 'theta': thr['theta'],
    }

    return {
        'num_criteria': n, 'num_alternatives': A, 'num_experts': E, 'num_objectives': nO,
        'criteria_names': criteria_names, 'criteria_types': criteria_types,
        'criteria': {i: criteria_names[i - 1] for i in I},
        'alternatives': alternatives, 'objectives_names': objectives_names,
        'objective_names': {o: objectives_names[o - 1] for o in O},
        'objective_definitions': {o: objectives_desc[o - 1] for o in O},
        'designated_objective': designated_obj,
        'I': I, 'O': O, 'active': active,
        'c_con': c_con, 'c_rng': c_rng, 'c': c, 'a': a, 'dr': dr, 'op': op, 'un': un, 'ce': ce,
        'ua': ua, 'q': q, 's': s_bar, 'o_i': o_i,
        'g': g, 'L': L, 'U': U, 'I_o': I_o, 'D': D, 'pairs': pairs,
        'obj_map': {o: [f"C{i}" for i in I if g[(i, o)] == 1] for o in O},
        'thresholds': thresholds, 'omega': omega, 'zeta': zeta,
        'rho_LB': rho_LB, 'rho_UB': rho_UB, 'M': M_big, 'epsilon': eps,
        'n_mc': n_mc, 'seed': seed, 'ties': ties,
        'raw': {'concern': concern, 'range': rng, 'a': a_raw, 'dr': dr_raw, 'op': op_raw, 'un': un_raw,
                'ce': ce_raw, 'ua': ua_raw, 'b': b_raw, 'o': o_votes},
    }


# ================================================================
# GATES, REWARD COEFFICIENTS, AND PORTFOLIO EVALUATION
# ================================================================

def gate_results(d):
    t = d['thresholds']; act = d['active']
    return {
        i: {
            "Completeness (concern)": (not act[1]) or d['c_con'][i] >= t['alpha_con'],
            "Completeness (range)": (not act[1]) or d['c_rng'][i] >= t['alpha_rng'],
            "Alignment": (not act[2]) or d['a'][i] >= t['lambda'],
            "Directness": (not act[3]) or d['dr'][i] >= t['psi'],
            "Operationality": (not act[7]) or d['op'][i] >= t['gamma'],
            "Understandability": (not act[8]) or d['un'][i] >= t['eta'],
            "Cost-effectiveness": (not act[9]) or d['ce'][i] >= t['tau'],
            "Unambiguity": (not act[10]) or d['ua'][i] >= t['mu'],
            "Monotone Coherence": (not act[11]) or d['q'][i] == 1,
            "Sensitivity": (not act[13]) or d['s'][i] >= t['theta'],
        }
        for i in d['I']
    }


REWARD_VECTORS = {
    'w1': (1, 'Completeness', 'c'), 'w2': (2, 'Alignment', 'a'), 'w3': (3, 'Directness', 'dr'),
    'w7': (7, 'Operationality', 'op'), 'w8': (8, 'Understandability', 'un'),
    'w9': (9, 'Cost-effectiveness', 'ce'), 'w10': (10, 'Unambiguity', 'ua'), 'w13': (13, 'Sensitivity', 's'),
}


def normalized_reward_coefficients(d, w):
    coeff = {i: 0.0 for i in d['I']}
    denominators = {}
    for key, (prop, name, vec) in REWARD_VECTORS.items():
        values = d[vec]
        weight = w[key] if d['active'][prop] else 0.0
        denominator = float(sum(values.values()))
        denominators[name] = denominator
        if denominator > 0:
            for i in d['I']:
                coeff[i] += weight * values[i] / denominator
    return coeff, denominators


def evaluate_portfolio(selected, d, w, reward_coeff):
    chosen = set(selected)
    N = len(chosen)
    if N == 0:
        return None
    n_o = {o: sum(d['g'][(i, o)] for i in chosen) for o in d['O']}
    if d['active'][4] and any(n_o[o] < 1 for o in d['O']):
        return None
    if d['active'][6]:
        quant = sum(d['o_i'][i] for i in chosen)
        if quant + TOL < d['rho_LB'] * N or quant - TOL > d['rho_UB'] * N:
            return None
    delta = d['thresholds']['delta']
    if d['active'][12] and any(i in chosen and k in chosen for (i, k), r in d['pairs'].items() if r > delta):
        return None

    d1_minus, d1_plus = max(d['omega'] - N, 0), max(N - d['omega'], 0)
    d2_minus, d2_plus = max(d['zeta'] - N, 0), max(N - d['zeta'], 0)
    rep_devs = {o: {"d_o1_minus": max(d['L'][o] - n_o[o], 0), "d_o1_plus": max(n_o[o] - d['L'][o], 0),
                    "d_o2_minus": max(d['U'][o] - n_o[o], 0), "d_o2_plus": max(n_o[o] - d['U'][o], 0)} for o in d['O']}

    reward = sum(reward_coeff[i] for i in chosen)
    nO = len(d['O'])
    rep_under_raw = (sum(rep_devs[o]["d_o1_minus"] / d['L'][o] for o in d['O'] if d['L'][o] > 0) / nO) if nO else 0.0
    rep_over_raw = (sum(rep_devs[o]["d_o2_plus"] / d['D'][o] for o in d['O'] if d['D'][o] > 0) / nO) if nO else 0.0
    rep_penalty = (w['w4_minus'] * rep_under_raw + w['w4_plus'] * rep_over_raw) if d['active'][4] else 0.0

    par_under_raw = d1_minus / d['omega'] if d['omega'] > 0 else 0.0
    upper_width = len(d['I']) - d['zeta']
    par_over_raw = d2_plus / upper_width if upper_width > 0 else 0.0
    parsimony_penalty = (w['w5_minus'] * par_under_raw + w['w5_plus'] * par_over_raw) if d['active'][5] else 0.0

    red_den = sum(d['pairs'].values())
    red_num = sum(r for (i, k), r in d['pairs'].items() if i in chosen and k in chosen)
    red_raw = red_num / red_den if red_den > 0 else 0.0
    distinctiveness_penalty = w['w12'] * red_raw if d['active'][12] else 0.0

    objective = reward - rep_penalty - parsimony_penalty - distinctiveness_penalty
    rho = sum(d['o_i'][i] for i in chosen) / N
    return {
        "selected": tuple(sorted(selected)), "N": N, "n_o": n_o, "rho": rho, "reward": reward,
        "rep_under_raw": rep_under_raw, "rep_over_raw": rep_over_raw, "rep_penalty": rep_penalty,
        "par_under_raw": par_under_raw, "par_over_raw": par_over_raw, "parsimony_penalty": parsimony_penalty,
        "red_num": red_num, "red_den": red_den, "red_raw": red_raw,
        "distinctiveness_penalty": distinctiveness_penalty, "objective": objective,
        "d1_minus": d1_minus, "d1_plus": d1_plus, "d2_minus": d2_minus, "d2_plus": d2_plus, "rep_devs": rep_devs,
    }


# ================================================================
# MILP MODEL (Pyomo) AND SOLVER
# ================================================================

def build_mcdm_model(d, w, reward_coeff):
    I, O, P = d['I'], d['O'], sorted(d['pairs'])
    M_big, eps = d['M'], d['epsilon']
    act = d['active']; t = d['thresholds']

    m = pyo.ConcreteModel(name="CREST_13_Property_Selection")
    m.I = pyo.Set(initialize=I, ordered=True)
    m.O = pyo.Set(initialize=O, ordered=True)
    m.P = pyo.Set(initialize=P, dimen=2, ordered=True)
    m.x = pyo.Var(m.I, domain=pyo.Binary)
    for nm in ["y_con", "y_rng", "y_a", "y_dr", "y_op", "y_un", "y_ce", "y_ua", "y_s"]:
        setattr(m, nm, pyo.Var(m.I, domain=pyo.Binary))
    m.h = pyo.Var(m.P, domain=pyo.Binary)
    m.t = pyo.Var(m.P, domain=pyo.Binary)
    m.N = pyo.Var(domain=pyo.NonNegativeIntegers)
    m.n = pyo.Var(m.O, domain=pyo.NonNegativeIntegers)
    m.d1_minus = pyo.Var(domain=pyo.NonNegativeIntegers)
    m.d1_plus = pyo.Var(domain=pyo.NonNegativeIntegers)
    m.d2_minus = pyo.Var(domain=pyo.NonNegativeIntegers)
    m.d2_plus = pyo.Var(domain=pyo.NonNegativeIntegers)
    m.do1_minus = pyo.Var(m.O, domain=pyo.NonNegativeIntegers)
    m.do1_plus = pyo.Var(m.O, domain=pyo.NonNegativeIntegers)
    m.do2_minus = pyo.Var(m.O, domain=pyo.NonNegativeIntegers)
    m.do2_plus = pyo.Var(m.O, domain=pyo.NonNegativeIntegers)

    def add_gate(prefix, values, threshold, y):
        setattr(m, f"{prefix}_upper", pyo.Constraint(m.I, rule=lambda mm, i: values[i] - threshold <= M_big * y[i] - eps))
        setattr(m, f"{prefix}_lower", pyo.Constraint(m.I, rule=lambda mm, i: values[i] - threshold >= -M_big * (1 - y[i]) - eps))
        setattr(m, f"{prefix}_select", pyo.Constraint(m.I, rule=lambda mm, i: mm.x[i] <= y[i]))

    if act[1]:
        add_gate("completeness_con", d['c_con'], t['alpha_con'], m.y_con)
        add_gate("completeness_rng", d['c_rng'], t['alpha_rng'], m.y_rng)
    if act[2]:
        add_gate("alignment", d['a'], t['lambda'], m.y_a)
    if act[3]:
        add_gate("directness", d['dr'], t['psi'], m.y_dr)
    if act[7]:
        add_gate("operationality", d['op'], t['gamma'], m.y_op)
    if act[8]:
        add_gate("understandability", d['un'], t['eta'], m.y_un)
    if act[9]:
        add_gate("cost_effectiveness", d['ce'], t['tau'], m.y_ce)
    if act[10]:
        add_gate("unambiguity", d['ua'], t['mu'], m.y_ua)
    if act[13]:
        add_gate("sensitivity", d['s'], t['theta'], m.y_s)

    m.N_definition = pyo.Constraint(expr=m.N == sum(m.x[i] for i in m.I))
    m.rep_count = pyo.Constraint(m.O, rule=lambda mm, o: mm.n[o] == sum(d['g'][(i, o)] * mm.x[i] for i in mm.I))
    if act[4]:
        m.rep_coverage = pyo.Constraint(m.O, rule=lambda mm, o: mm.n[o] >= 1)
    m.rep_lower = pyo.Constraint(m.O, rule=lambda mm, o: mm.n[o] + mm.do1_minus[o] - mm.do1_plus[o] == d['L'][o])
    m.rep_upper = pyo.Constraint(m.O, rule=lambda mm, o: mm.n[o] + mm.do2_minus[o] - mm.do2_plus[o] == d['U'][o])
    m.parsimony_lower = pyo.Constraint(expr=m.N + m.d1_minus - m.d1_plus == d['omega'])
    m.parsimony_upper = pyo.Constraint(expr=m.N + m.d2_minus - m.d2_plus == d['zeta'])

    if act[6]:
        m.assessment_lower = pyo.Constraint(expr=d['rho_LB'] * m.N <= sum(d['o_i'][i] * m.x[i] for i in m.I))
        m.assessment_upper = pyo.Constraint(expr=sum(d['o_i'][i] * m.x[i] for i in m.I) <= d['rho_UB'] * m.N)
    if act[11]:
        m.monotone = pyo.Constraint(m.I, rule=lambda mm, i: mm.x[i] <= d['q'][i])
    if act[12]:
        delta = t['delta']
        m.dist_upper = pyo.Constraint(m.P, rule=lambda mm, i, k: d['pairs'][(i, k)] - delta <= M_big * mm.h[i, k])
        m.dist_lower = pyo.Constraint(m.P, rule=lambda mm, i, k: d['pairs'][(i, k)] - delta >= eps - M_big * (1 - mm.h[i, k]))
        m.dist_select = pyo.Constraint(m.P, rule=lambda mm, i, k: mm.x[i] + mm.x[k] <= 2 - mm.h[i, k])
        m.lin1 = pyo.Constraint(m.P, rule=lambda mm, i, k: mm.t[i, k] <= mm.x[i])
        m.lin2 = pyo.Constraint(m.P, rule=lambda mm, i, k: mm.t[i, k] <= mm.x[k])
        m.lin3 = pyo.Constraint(m.P, rule=lambda mm, i, k: mm.t[i, k] >= mm.x[i] + mm.x[k] - 1)

    rep_penalty = 0.0
    if act[4]:
        rep_penalty = sum(w['w4_minus'] * m.do1_minus[o] / d['L'][o] + w['w4_plus'] * m.do2_plus[o] / d['D'][o] for o in m.O) / len(O)
    parsimony_penalty = 0.0
    if act[5]:
        if d['omega'] > 0:
            parsimony_penalty += w['w5_minus'] * m.d1_minus / d['omega']
        if len(I) > d['zeta']:
            parsimony_penalty += w['w5_plus'] * m.d2_plus / (len(I) - d['zeta'])
    red_den = sum(d['pairs'].values())
    distinctiveness_penalty = (w['w12'] * sum(d['pairs'][(i, k)] * m.t[i, k] for i, k in m.P) / red_den) if (act[12] and red_den > 0) else 0.0
    reward = sum(reward_coeff[i] * m.x[i] for i in m.I)
    m.objective = pyo.Objective(expr=reward - rep_penalty - parsimony_penalty - distinctiveness_penalty, sense=pyo.maximize)
    return m


def pick_solver():
    for name in ("highs", "appsi_highs", "cbc", "glpk"):
        try:
            s = SolverFactory(name)
            if s.available(False):
                return s, name
        except Exception:
            continue
    raise RuntimeError("No MILP solver found (tried HiGHS, CBC, GLPK).")


def solve_by_enumeration(d, w, reward_coeff, gates):
    eligible = [i for i in d['I'] if all(gates[i].values())]
    best = None
    for size in range(1, len(eligible) + 1):
        for selected in combinations(eligible, size):
            cand = evaluate_portfolio(selected, d, w, reward_coeff)
            if cand is None:
                continue
            key = (cand['objective'], tuple(-i for i in cand['selected']))
            if best is None or key > (best['objective'], tuple(-i for i in best['selected'])):
                best = cand
    if best is None:
        raise RuntimeError("No feasible portfolio satisfies the active properties.")
    return best


def solve_model(d, w):
    """Solve the MILP. Returns (solution, gates, reward_coeff, denominators, method)."""
    gates = gate_results(d)
    reward_coeff, denominators = normalized_reward_coefficients(d, w)
    try:
        model = build_mcdm_model(d, w, reward_coeff)
        solver, name = pick_solver()
        result = solver.solve(model, tee=False)
        if result.solver.termination_condition != TerminationCondition.optimal:
            raise RuntimeError(f"MILP did not reach optimality: {result.solver.termination_condition}")
        selected = tuple(i for i in d['I'] if pyo.value(model.x[i]) > 0.5)
        solution = evaluate_portfolio(selected, d, w, reward_coeff)
        if solution is None:
            raise RuntimeError("Solver returned a portfolio that failed post-solve validation.")
        return solution, gates, reward_coeff, denominators, f"Pyomo + {name.upper()} MILP"
    except RuntimeError as exc:
        if "did not reach optimality" in str(exc) and "infeasible" in str(exc).lower():
            raise
        if len(d['I']) > 20:
            raise
        solution = solve_by_enumeration(d, w, reward_coeff, gates)
        return solution, gates, reward_coeff, denominators, f"Exact enumeration (fallback: {exc})"


# ================================================================
# RESULT FRAMES AND EXCEL EXPORT
# ================================================================

def build_result_frames(d, w, solution, gates, reward_coeff, denominators, method):
    chosen = set(solution['selected'])
    delta = d['thresholds']['delta']
    act = d['active']

    def status(p):
        return "Active" if act[p] else "Inactive"

    def gate_display(p, passed):
        return ("Pass" if passed else "Fail") if act[p] else "Inactive"

    def eff_w(p, key):
        return w[key] if act[p] else 0.0

    rows = []
    for i in d['I']:
        failures = [nm for nm, ok in gates[i].items() if not ok]
        conflicts = []
        if act[12]:
            for (j, k), r in d['pairs'].items():
                if r <= delta:
                    continue
                other = k if j == i else j if k == i else None
                if other in chosen:
                    conflicts.append(f"C{other} {d['criteria'][other]}")
        if i in chosen:
            final, reason = "Selected", "None"
        elif failures:
            final, reason = "Filtered before portfolio optimization", "; ".join(failures)
        elif conflicts:
            final, reason = "Excluded by Distinctiveness conflict", "; ".join(conflicts)
        else:
            final, reason = "Eligible but not selected by optimization", "Portfolio-level objective trade-off"
        represented = [f"O{o}" for o in d['O'] if d['g'][(i, o)] == 1]
        rows.append({
            "ID": f"C{i}", "Criterion": d['criteria'][i], "Type": d['criteria_types'][i - 1], "x_i": int(i in chosen),
            "Final status": final, "Primary exclusion reason": reason,
            "All failed gates": "; ".join(failures) if failures else "None",
            "Distinctiveness conflicts with selected": "; ".join(conflicts) if conflicts else "None",
            "Represented objectives": ", ".join(represented) if represented else "None",
            "o_i": d['o_i'][i], "Assessment Mode": "Quantitative" if d['o_i'][i] == 1 else "Qualitative or ordinal",
            "Normalized reward contribution": reward_coeff[i] if i in chosen else 0.0,
            "c_con": d['c_con'][i], "c_rng": d['c_rng'][i], "c (min)": d['c'][i], "a": d['a'][i], "dr": d['dr'][i],
            "op": d['op'][i], "un": d['un'][i], "ce": d['ce'][i], "ua": d['ua'][i], "q": d['q'][i], "s_bar": d['s'][i],
            "Completeness concern": gate_display(1, gates[i]["Completeness (concern)"]),
            "Completeness range": gate_display(1, gates[i]["Completeness (range)"]),
            "Alignment": gate_display(2, gates[i]["Alignment"]),
            "Directness": gate_display(3, gates[i]["Directness"]),
            "Operationality": gate_display(7, gates[i]["Operationality"]),
            "Understandability": gate_display(8, gates[i]["Understandability"]),
            "Cost-effectiveness": gate_display(9, gates[i]["Cost-effectiveness"]),
            "Unambiguity": gate_display(10, gates[i]["Unambiguity"]),
            "Monotone Coherence": gate_display(11, gates[i]["Monotone Coherence"]),
            "Sensitivity": gate_display(13, gates[i]["Sensitivity"]),
        })
    decisions = pd.DataFrame(rows)

    selected = decisions.loc[decisions["x_i"].eq(1), ["ID", "Criterion", "Type", "Represented objectives", "Normalized reward contribution"]].copy()
    selected.insert(0, "Portfolio order", range(1, len(selected) + 1))

    cov_rows = []
    for o in d['O']:
        sel_i = [i for i in solution['selected'] if d['g'][(i, o)] == 1]
        dev = solution['rep_devs'][o]
        cov_rows.append({
            "Objective": f"O{o}", "Name": d['objective_names'][o], "Definition": d['objective_definitions'][o],
            "|I_o|": d['I_o'][o], "L_o": d['L'][o], "U_o": d['U'][o], "D_o": d['D'][o], "n_o": solution['n_o'][o],
            "Hard coverage": (("Met" if solution['n_o'][o] >= 1 else "Not met") if act[4] else "Inactive"),
            "Soft target status": (("Within target" if d['L'][o] <= solution['n_o'][o] <= d['U'][o] else "Outside target") if act[4] else "Inactive"),
            "d_o1^-": dev["d_o1_minus"], "d_o1^+": dev["d_o1_plus"], "d_o2^-": dev["d_o2_minus"], "d_o2^+": dev["d_o2_plus"],
            "Selected representatives": "; ".join(f"C{i} {d['criteria'][i]}" for i in sel_i),
        })
    coverage = pd.DataFrame(cov_rows)

    pair_rows = []
    for (i, k), r in sorted(d['pairs'].items()):
        both, flagged = (i in chosen and k in chosen), r > delta
        pair_rows.append({
            "i": f"C{i}", "Criterion i": d['criteria'][i], "k": f"C{k}", "Criterion k": d['criteria'][k],
            "tilde_r_ik": r, "delta": delta, "h_ik": int(flagged), "t_ik": int(both),
            "Jointly selected": "Yes" if both else "No",
            "Status": (("Prohibited pair" if flagged else "Within cap") if act[12] else "Property inactive"),
            "Objective numerator contribution": r if (both and act[12]) else 0.0,
        })
    pairwise = pd.DataFrame(pair_rows)

    def reward_row(key):
        p, name, vec = REWARD_VECTORS[key]
        coef = eff_w(p, key)
        den = denominators[name]
        raw = sum(d[vec][i] for i in chosen) / den if den > 0 else 0.0
        return {"Property no.": p, "Component": name, "Status": status(p), "Notation": WEIGHT_NOTATION[key], "Type": "Reward",
                "Raw normalized value": raw, "Coefficient used": coef, "Signed contribution to Z": coef * raw}

    def penalty_row(p, comp, key, raw):
        coef = eff_w(p, key)
        return {"Property no.": p, "Component": comp, "Status": status(p), "Notation": WEIGHT_NOTATION[key], "Type": "Penalty",
                "Raw normalized value": raw, "Coefficient used": coef, "Signed contribution to Z": -coef * raw}

    br = [reward_row(k) for k in ('w1', 'w2', 'w3')]
    br += [
        penalty_row(4, "Representativeness (under)", 'w4_minus', solution['rep_under_raw']),
        penalty_row(4, "Representativeness (over)", 'w4_plus', solution['rep_over_raw']),
        penalty_row(5, "Parsimony (under-complexity)", 'w5_minus', solution['par_under_raw']),
        penalty_row(5, "Parsimony (over-complexity)", 'w5_plus', solution['par_over_raw']),
        {"Property no.": 6, "Component": "Assessment Mode", "Status": status(6), "Notation": "rho", "Type": "Composition control",
         "Raw normalized value": solution['rho'], "Coefficient used": 0.0, "Signed contribution to Z": 0.0},
    ]
    br += [reward_row(k) for k in ('w7', 'w8', 'w9', 'w10')]
    br.append({"Property no.": 11, "Component": "Monotone Coherence", "Status": status(11), "Notation": "q_i", "Type": "Hard veto",
               "Raw normalized value": min(d['q'][i] for i in chosen), "Coefficient used": 0.0, "Signed contribution to Z": 0.0})
    br.append(penalty_row(12, "Distinctiveness overlap", 'w12', solution['red_raw']))
    br.append(reward_row('w13'))
    breakdown = pd.DataFrame(br)

    active_nums = [p for p in range(1, 14) if act[p]]
    inactive_nums = [p for p in range(1, 14) if not act[p]]
    summary = pd.DataFrame([
        ["Status", "Optimal", "Solved under the properties activated in the template"],
        ["Solution method", method, ""],
        ["Objective value Z", solution['objective'], "Only active weighted properties contribute"],
        ["Selected criteria N", solution['N'], (f"Soft target range: [{d['omega']}, {d['zeta']}]" if act[5] else "Property V inactive")],
        ["Selected portfolio", ", ".join(f"C{i}" for i in solution['selected']), "; ".join(d['criteria'][i] for i in solution['selected'])],
        ["Filtered by active score/veto gates", int(decisions["Final status"].eq("Filtered before portfolio optimization").sum()), "See Criterion_Decisions"],
        ["Eligible but not selected", int(decisions["Final status"].eq("Eligible but not selected by optimization").sum()), "Portfolio-level trade-off"],
        ["Active properties", len(active_nums), ", ".join(f"{p} {PROPERTIES[p]}" for p in active_nums) or "None"],
        ["Inactive properties", len(inactive_nums), ", ".join(f"{p} {PROPERTIES[p]}" for p in inactive_nums) or "None"],
        ["Assessment Mode", status(6), "Composition interval imposed" if act[6] else "Composition interval not imposed"],
        ["Reported quantitative proportion rho", solution['rho'], "Computed after selection from o_i tags"],
        ["Parsimony deviations", f"d1-={solution['d1_minus']}; d1+={solution['d1_plus']}; d2-={solution['d2_minus']}; d2+={solution['d2_plus']}", "Penalized only when Property V is active"],
        ["Weighting", "SWING (app Step 3)", "Weights of inactive properties are locked to 0; active weights sum to 1"],
    ], columns=["Metric", "Value", "Interpretation"])

    activation = pd.DataFrame([[p, PROPERTIES[p], "Yes" if act[p] else "No", PROPERTY_EFFECTS[p]] for p in range(1, 14)],
                              columns=["Property no.", "Property", "Active?", "Model effect when active"])

    def use(p, role):
        return f"{'Applied' if act[p] else 'Ignored'}: {role}"

    t = d['thresholds']
    controls = pd.DataFrame([
        [1, "Threshold", "alpha^con", t['alpha_con'], use(1, "concern gate")],
        [1, "Threshold", "alpha^rng", t['alpha_rng'], use(1, "range gate")],
        [1, "Weight", "w1", w['w1'], use(1, "reward")],
        [2, "Threshold", "lambda", t['lambda'], use(2, "gate")],
        [2, "Weight", "w2", w['w2'], use(2, "reward")],
        [3, "Threshold", "psi", t['psi'], use(3, "gate")],
        [3, "Weight", "w3", w['w3'], use(3, "reward")],
        [4, "Weight", "w4^-", w['w4_minus'], use(4, "under-representation penalty")],
        [4, "Weight", "w4^+", w['w4_plus'], use(4, "over-representation penalty")],
        [5, "Target", "omega", d['omega'], use(5, "soft lower target")],
        [5, "Target", "zeta", d['zeta'], use(5, "soft upper target")],
        [5, "Weight", "w5^-", w['w5_minus'], use(5, "under-complexity penalty")],
        [5, "Weight", "w5^+", w['w5_plus'], use(5, "over-complexity penalty")],
        [6, "Composition control", "rho^LB", d['rho_LB'], use(6, "lower composition bound")],
        [6, "Composition control", "rho^UB", d['rho_UB'], use(6, "upper composition bound")],
        [7, "Threshold", "gamma", t['gamma'], use(7, "gate")],
        [7, "Weight", "w7", w['w7'], use(7, "reward")],
        [8, "Threshold", "eta", t['eta'], use(8, "gate")],
        [8, "Weight", "w8", w['w8'], use(8, "reward")],
        [9, "Threshold", "tau", t['tau'], use(9, "gate")],
        [9, "Weight", "w9", w['w9'], use(9, "reward")],
        [10, "Threshold", "mu", t['mu'], use(10, "gate")],
        [10, "Weight", "w10", w['w10'], use(10, "reward")],
        [11, "Criterion tag", "q_i", "See XI sheet", use(11, "hard veto")],
        [12, "Threshold", "delta", t['delta'], use(12, "pairwise exclusion")],
        [12, "Weight", "w12", w['w12'], use(12, "overlap penalty")],
        [13, "Threshold", "theta", t['theta'], use(13, "gate")],
        [13, "Weight", "w13", w['w13'], use(13, "reward")],
        [13, "Monte Carlo", "N_MC", d['n_mc'], use(13, "runs")],
        [13, "Monte Carlo", "seed", d['seed'], use(13, "random seed")],
        ["", "MILP setting", "M", d['M'], "Big-M constant"],
        ["", "MILP setting", "epsilon", d['epsilon'], "Strict-comparison tolerance"],
    ], columns=["Property no.", "Type", "Notation", "Value", "Status / role"])

    return {
        "Summary": summary, "Property_Activation": activation, "Selected_Portfolio": selected,
        "Criterion_Decisions": decisions, "Objective_Coverage": coverage,
        "Objective_Breakdown": breakdown, "Pairwise_Review": pairwise, "Model_Controls": controls,
    }


def export_results_excel(frames):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet[:31], index=False, startrow=3)
    buffer.seek(0)
    wb = openpyxl.load_workbook(buffer)
    navy, blue, pale, line = "17365D", "1F4E78", "D9EAF7", "B7C9D6"
    thin = Side(style="thin", color=line)
    for ws in wb.worksheets:
        max_col = max(ws.max_column, 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.cell(1, 1, ws.title.replace("_", " "))
        ws.cell(1, 1).fill = PatternFill("solid", fgColor=navy)
        ws.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=15)
        ws.row_dimensions[1].height = 25
        for cell in ws[4]:
            cell.fill = PatternFill("solid", fgColor=blue)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[4].height = 30
        for row in ws.iter_rows(min_row=5):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if row[0].row % 2:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=pale)
        ws.freeze_panes = "A5"
        for col in range(1, max_col + 1):
            values = [str(ws.cell(r, col).value or "") for r in range(4, min(ws.max_row, 60) + 1)]
            ws.column_dimensions[ws.cell(4, col).column_letter].width = min(48, max(11, max(map(len, values), default=10) + 2))
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ================================================================
# UI HELPER FUNCTIONS
# ================================================================

def show_progress_indicator(current_step):
    steps = [("1", "Generate Template", "📝"), ("2", "Upload & Extract", "📤"),
             ("3", "Set Weights", "⚖️"), ("4", "Run Optimization", "🚀")]
    cols = st.columns(4)
    for idx, (step_num, step_name, icon) in enumerate(steps):
        with cols[idx]:
            if idx + 1 < current_step:
                bg, color, content = "#10b981", "#10b981", "✓"
            elif idx + 1 == current_step:
                bg, color, content = "linear-gradient(135deg, #667eea 0%, #764ba2 100%)", "#667eea", icon
            else:
                bg, color, content = "#e5e7eb", "#6b7280", step_num
            st.markdown(f"""
            <div style="text-align: center;">
                <div style="width: 40px; height: 40px; border-radius: 50%; background: {bg}; color: white;
                            display: flex; align-items: center; justify-content: center;
                            margin: 0 auto 0.5rem auto; font-weight: 600;">{content}</div>
                <div style="font-size: 0.875rem; color: {color}; font-weight: 600;">{step_name}</div>
            </div>
            """, unsafe_allow_html=True)


def show_navigation_buttons(current_step):
    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        if current_step > 1:
            if st.button("⬅️ Back", use_container_width=True, type="secondary"):
                st.session_state.current_step = current_step - 1
                st.rerun()
    with col3:
        if current_step < 4:
            can_proceed = (current_step == 1) or (current_step == 2 and st.session_state.data is not None) \
                or (current_step == 3 and st.session_state.weights is not None)
            if st.button("Next ➡️", use_container_width=True, type="primary", disabled=not can_proceed):
                st.session_state.current_step = current_step + 1
                st.rerun()


# ================================================================
# STEP 1: GENERATE TEMPLATE
# ================================================================

def show_step1_generate_template():
    st.header("📝 Step 1: Generate Excel Template")
    st.markdown("Configure your problem parameters and generate a customized Excel template.")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Problem Structure")
        num_criteria = st.number_input("Number of Criteria", min_value=1, value=16, step=1, key="num_criteria")
        num_alternatives = st.number_input("Number of Alternatives", min_value=1, value=7, step=1, key="num_alt",
                                           help="Used only by Properties XII (Distinctiveness) and XIII (Sensitivity).")
        num_experts = st.number_input("Number of Experts", min_value=1, value=3, step=1, key="num_exp")
        num_objectives = st.number_input("Number of Objectives", min_value=1, value=7, step=1, key="num_obj")
    with col2:
        st.subheader("Parsimony Targets (Property V)")
        omega = st.number_input("Target Minimum (omega)", min_value=0, value=5, step=1, key="omega")
        zeta = st.number_input("Target Maximum (zeta)", min_value=1, value=9, step=1, key="zeta")
        st.info("💡 Soft target range for the number of selected criteria. Deviations are penalized, not forbidden.")

    with st.expander("🔘 Property Activation (default: all active)"):
        st.markdown("Uncheck a property to exclude it. Its sheet is prefilled with 0, it imposes no gate or penalty, and its weight is locked to 0 in Step 3.")
        active = {}
        cols = st.columns(3)
        for p in range(1, 14):
            with cols[(p - 1) % 3]:
                active[p] = st.checkbox(f"{p}. {PROPERTIES[p]}", value=True, key=f"active_{p}", help=PROPERTY_EFFECTS[p])

    with st.expander("🎯 Objective Representation Targets L(o) and U(o) (Property IV)"):
        st.markdown("Desired lower and upper number of selected criteria representing each objective (soft targets). Defaults: L(o) = 1, U(o) = 2.")
        L_list, U_list = [], []
        cols = st.columns(2)
        for o in range(int(num_objectives)):
            with cols[o % 2]:
                c1, c2 = st.columns(2)
                L_list.append(c1.number_input(f"O{o+1}: L(o)", min_value=1, value=1, step=1, key=f"L_{o}"))
                U_list.append(c2.number_input(f"O{o+1}: U(o)", min_value=1, value=2, step=1, key=f"U_{o}"))

    with st.expander("⚙️ Thresholds and Bounds"):
        c1, c2, c3 = st.columns(3)
        with c1:
            alpha_con = st.number_input("I. Completeness, concern (alpha^con)", value=6.0, key="alpha_con")
            alpha_rng = st.number_input("I. Completeness, range (alpha^rng)", value=6.0, key="alpha_rng")
            lambda_th = st.number_input("II. Alignment (lambda)", value=6.5, key="lambda")
            psi = st.number_input("III. Directness (psi)", value=6.5, key="psi")
        with c2:
            rho_LB = st.number_input("VI. Assessment Mode lower bound (rho^LB)", min_value=0.0, max_value=1.0, value=0.0, key="rho_lb")
            rho_UB = st.number_input("VI. Assessment Mode upper bound (rho^UB)", min_value=0.0, max_value=1.0, value=1.0, key="rho_ub")
            gamma = st.number_input("VII. Operationality (gamma)", value=6.0, key="gamma")
            eta = st.number_input("VIII. Understandability (eta)", value=6.5, key="eta")
        with c3:
            tau = st.number_input("IX. Cost-effectiveness (tau)", value=7.0, key="tau")
            mu = st.number_input("X. Unambiguity (mu)", value=7.0, key="mu")
            delta = st.number_input("XII. Distinctiveness (delta)", value=0.75, key="delta")
            theta = st.number_input("XIII. Sensitivity (theta)", value=0.035, format="%.4f", key="theta")

    with st.expander("🧪 Advanced Settings (Monte Carlo and MILP constants)"):
        c1, c2, c3, c4 = st.columns(4)
        n_mc = c1.number_input("Sensitivity Monte Carlo runs (N_MC)", min_value=10, value=1000, step=100, key="n_mc")
        seed = c2.number_input("Random seed", min_value=0, value=42, step=1, key="seed")
        M_big = c3.number_input("Big-M constant (M)", min_value=1.0, value=10000.0, key="M_big")
        eps = c4.number_input("Epsilon", min_value=1e-9, value=1e-6, format="%.1e", key="eps")

    st.markdown("<br>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("🎨 Generate Excel Template", type="primary", use_container_width=True):
            errors = []
            if int(omega) >= int(zeta):
                errors.append("omega must be strictly smaller than zeta.")
            if int(zeta) > int(num_criteria):
                errors.append("zeta cannot exceed the number of criteria.")
            if active[10] and int(num_experts) < 2:
                errors.append("Unambiguity (Property X) requires at least two experts. Add experts or deactivate Property X.")
            if any(L_list[o] > U_list[o] for o in range(int(num_objectives))):
                errors.append("Each objective must satisfy L(o) <= U(o).")
            if rho_LB > rho_UB:
                errors.append("rho^LB must not exceed rho^UB.")
            if errors:
                for e in errors:
                    st.error(f"❌ {e}")
                return
            with st.spinner("Generating template..."):
                try:
                    thresholds = {'alpha_con': alpha_con, 'alpha_rng': alpha_rng, 'lambda': lambda_th, 'psi': psi,
                                  'rho_LB': rho_LB, 'rho_UB': rho_UB, 'gamma': gamma, 'eta': eta, 'tau': tau,
                                  'mu': mu, 'delta': delta, 'theta': theta}
                    buffer, config = generate_excel_template(
                        int(num_criteria), int(num_alternatives), int(num_experts), int(num_objectives),
                        int(omega), int(zeta), [int(v) for v in L_list], [int(v) for v in U_list],
                        active, thresholds, int(n_mc), int(seed), float(M_big), float(eps))
                    st.session_state.config = config
                    st.success("✅ Template generated successfully!")
                    st.download_button(
                        label="📥 Download Excel Template", data=buffer,
                        file_name=f"CREST_Template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary")
                    inactive = [PROPERTIES[p] for p in range(1, 14) if not active[p]]
                    st.markdown(f"""
                    <div class="info-box">
                        <strong>📋 Next Steps:</strong><br>
                        1. Download the Excel template<br>
                        2. Fill in the yellow cells with expert data (criteria names and types, alternatives, objectives, ratings, decision matrices)<br>
                        3. Save the file<br>
                        4. Click "Next" to proceed to upload<br>
                        <strong>Inactive properties:</strong> {", ".join(inactive) if inactive else "None"}
                    </div>
                    """, unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")


# ================================================================
# STEP 2: UPLOAD & EXTRACT
# ================================================================

def show_step2_upload_extract():
    st.header("📤 Step 2: Upload Filled Template")
    st.markdown("Upload your completed Excel template. The app computes all property scores (medians, majority tags, correlations, Monte Carlo sensitivities) and applies the threshold gates.")

    uploaded_file = st.file_uploader("Choose Excel file", type=['xlsx'], key="upload")
    if not uploaded_file:
        return
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        run = st.button("🔍 Extract Data", type="primary", use_container_width=True)
    if run:
        with st.spinner("Reading Excel file and computing property scores..."):
            try:
                data = read_mcdm_template(uploaded_file)
                st.session_state.data = data
                st.session_state.weights = None
                st.session_state.solution = None
                st.session_state.result_frames = None
                # Reset SWING widget state so weights are rebuilt for the new data
                for k in list(st.session_state.keys()):
                    if k.startswith("slider_") or k.startswith("input_"):
                        del st.session_state[k]
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
                return

    data = st.session_state.data
    if data is None:
        return
    st.markdown('<div class="success-box"><strong>✅ Data extracted successfully!</strong></div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Criteria", data['num_criteria'])
    c2.metric("Alternatives", data['num_alternatives'])
    c3.metric("Experts", data['num_experts'])
    c4.metric("Objectives", data['num_objectives'])
    if data['ties']:
        st.warning("Assessment Mode is inactive, but the strict-majority vote is tied for: " + "; ".join(data['ties']) + ". These are tagged qualitative (0) for reporting rho only.")

    inactive = [f"{p}. {PROPERTIES[p]}" for p in range(1, 14) if not data['active'][p]]
    if inactive:
        st.info("Inactive properties (ignored, weight locked to 0): " + ", ".join(inactive))

    gates = gate_results(data)
    t = data['thresholds']
    with st.expander("📋 Property scores and threshold gates", expanded=True):
        rows = []
        for i in data['I']:
            failed = [nm for nm, ok in gates[i].items() if not ok]
            rows.append({
                'ID': f"C{i}", 'Name': data['criteria'][i], 'Type': data['criteria_types'][i - 1],
                'c_con': data['c_con'][i], 'c_rng': data['c_rng'][i], 'c': data['c'][i],
                'a': data['a'][i], 'dr': data['dr'][i], 'op': data['op'][i], 'un': data['un'][i],
                'ce': data['ce'][i], 'ua': data['ua'][i], 'q': data['q'][i], 's_bar': round(data['s'][i], 4),
                'o_i': data['o_i'][i], 'Gate status': "Eligible" if not failed else "Fails: " + "; ".join(failed),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        st.caption(f"Thresholds: alpha^con={t['alpha_con']}, alpha^rng={t['alpha_rng']}, lambda={t['lambda']}, psi={t['psi']}, gamma={t['gamma']}, eta={t['eta']}, tau={t['tau']}, mu={t['mu']}, delta={t['delta']}, theta={t['theta']}")

    with st.expander("🎯 Objectives and consolidated assignments (strict majority)"):
        for o in data['O']:
            st.write(f"**O{o}: {data['objective_names'][o]}** (L={data['L'][o]}, U={data['U'][o]}, |I_o|={data['I_o'][o]}) → Criteria: {data['obj_map'].get(o, [])}")

    if data['active'][12]:
        with st.expander("🔗 Pooled absolute correlations (Distinctiveness)"):
            n = data['num_criteria']
            mat = np.eye(n)
            for (i, k), r in data['pairs'].items():
                mat[i - 1, k - 1] = mat[k - 1, i - 1] = r
            labels = [f"C{i}" for i in data['I']]
            st.dataframe(pd.DataFrame(mat, index=labels, columns=labels).round(3), use_container_width=True)
            flagged = [(i, k, r) for (i, k), r in data['pairs'].items() if r > t['delta']]
            if flagged:
                st.warning("Pairs exceeding delta (cannot be selected jointly): " + "; ".join(f"C{i}-C{k} ({r:.3f})" for i, k, r in flagged))
            else:
                st.success(f"No pair exceeds delta = {t['delta']}.")

    st.info("✅ Ready! Click 'Next' to set weights.")


# ================================================================
# STEP 3: SWING WEIGHTING
# ================================================================

def get_property_ranges(data):
    """Best and worst values of each objective term (all terms are normalized to [0, 1] in the model)."""
    ranges = {}
    for key, (p, name, vec) in REWARD_VECTORS.items():
        vals = list(data[vec].values())
        ranges[key] = {'best': max(vals) if vals else 0.0, 'worst': min(vals) if vals else 0.0, 'higher_is_better': True,
                       'label': 'score'}
    for key in ('w4_minus', 'w4_plus', 'w5_minus', 'w5_plus', 'w12'):
        ranges[key] = {'best': 0.0, 'worst': 1.0, 'higher_is_better': False, 'label': 'normalized penalty'}
    return ranges


def show_step3_set_weights():
    st.header("⚖️ Step 3: SWING Weighting")
    data = st.session_state.data
    if not data:
        st.warning("⚠️ Please upload and extract data first!")
        return
    active = data['active']
    ranges = get_property_ranges(data)

    st.markdown("""
    <div class="info-box">
        <strong>💡 SWING weighting of the property-based objective terms:</strong><br>
        1. Select the active term whose swing from worst to best level is the most valuable; it receives a score of 100<br>
        2. Rate every other active term relative to it (0 = no value, 50 = half as valuable, 100 = equally valuable)<br>
        3. Scores are normalized so that the weights of the active terms sum to exactly 1.0<br>
        4. Terms of inactive properties are locked to 0 and cannot be changed<br><br>
        <strong>📊 Best/Worst Values Guide:</strong> reward terms show the range of property scores across criteria (higher is better);
        penalty terms show the normalized penalty range (0 = no penalty, 1 = maximal penalty).
    </div>
    """, unsafe_allow_html=True)

    active_keys = [k for k, (p, _, _, _) in WEIGHT_COMPONENTS.items() if active[p]]
    inactive_keys = [k for k in WEIGHT_COMPONENTS if k not in active_keys]
    if not active_keys:
        st.error("All weighted properties are inactive. Activate at least one weighted property in Step 1.")
        return
    if st.session_state.reference_component not in active_keys:
        st.session_state.reference_component = active_keys[0]

    col1, col2 = st.columns([2, 1])
    with col1:
        st.subheader("3.1. Select Reference Component")
        st.markdown("**Which active term has the MOST valuable swing from its worst to its best level?**")
        reference = st.selectbox(
            "Reference Component (automatically set to 100)", options=active_keys,
            format_func=lambda k: f"{WEIGHT_COMPONENTS[k][1]} ({WEIGHT_NOTATION[k]}) - {WEIGHT_COMPONENTS[k][3]}",
            key="ref_component_select", index=active_keys.index(st.session_state.reference_component))
        st.session_state.reference_component = reference
        rr = ranges[reference]
        arrow_b, arrow_w = ("↑", "↓") if rr['higher_is_better'] else ("↓", "↑")
        st.markdown(f"""
        <div class="reference-box">
            <strong>Reference: {WEIGHT_COMPONENTS[reference][1]} ({WEIGHT_NOTATION[reference]})</strong><br>
            This term is set to 100 (maximum value)<br>
            <span style="font-size: 0.875rem; color: #92400e;">Best: {rr['best']:.3f} {arrow_b} | Worst: {rr['worst']:.3f} {arrow_w}</span>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.subheader("3.2. Rate Other Components Relative to Reference")
        st.markdown("*Scale: 0 = no value, 50 = half as valuable, 100 = equally valuable*")

        def sync_from_slider(k):
            st.session_state[f"input_{k}"] = st.session_state[f"slider_{k}"]

        def sync_from_input(k):
            st.session_state[f"slider_{k}"] = st.session_state[f"input_{k}"]

        raw = {reference: 100.0}
        for key in active_keys:
            if key == reference:
                continue
            p, name, kind, desc = WEIGHT_COMPONENTS[key]
            sk, ik = f"slider_{key}", f"input_{key}"
            if sk not in st.session_state:
                st.session_state[sk] = 50.0
            if ik not in st.session_state:
                st.session_state[ik] = 50.0
            pr = ranges[key]
            col_range, col_slider, col_input = st.columns([1, 3, 1])
            with col_range:
                st.markdown("<br>", unsafe_allow_html=True)
                ab, aw = ("↑", "↓") if pr['higher_is_better'] else ("↓", "↑")
                st.markdown(f"""
                <div style="font-size: 0.75rem; padding: 0.25rem; background: #f0fdf4; border-radius: 4px; margin-bottom: 0.25rem;">
                    <span style="color: #10b981; font-weight: 600;">{ab} Best:</span> {pr['best']:.3f}</div>
                <div style="font-size: 0.75rem; padding: 0.25rem; background: #fef2f2; border-radius: 4px;">
                    <span style="color: #ef4444; font-weight: 600;">{aw} Worst:</span> {pr['worst']:.3f}</div>
                """, unsafe_allow_html=True)
            with col_slider:
                st.slider(f"**{name}** ({WEIGHT_NOTATION[key]}, {kind}) relative to {WEIGHT_COMPONENTS[reference][1]}",
                          min_value=0.0, max_value=100.0, step=1.0, key=sk, on_change=sync_from_slider, args=(key,),
                          help=f"{desc}\n\n100 = As valuable as the reference\n50 = Half as valuable\n0 = Not valuable")
            with col_input:
                st.markdown("<br>", unsafe_allow_html=True)
                st.number_input("Precise value", min_value=0.0, max_value=100.0, step=0.01, format="%.2f",
                                key=ik, on_change=sync_from_input, args=(key,), label_visibility="collapsed")
            raw[key] = float(st.session_state[sk])

        if inactive_keys:
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("**Locked terms (inactive properties, weight = 0):**")
            st.markdown(", ".join(f"{WEIGHT_COMPONENTS[k][1]} ({WEIGHT_NOTATION[k]})" for k in inactive_keys))

    total = sum(raw.values())
    normalized = {k: (raw[k] / total if k in raw else 0.0) for k in WEIGHT_COMPONENTS}
    st.session_state.weights = normalized

    with col2:
        st.subheader("Normalized Weights")
        st.markdown(f"""
        <div style="background: #fef3c7; padding: 0.75rem; border-radius: 8px; margin-bottom: 1rem; border: 2px solid #f59e0b;">
            <div style="font-weight: 600; color: #92400e;">Reference Component:</div>
            <div style="font-size: 1.1rem; font-weight: 700; color: #92400e;">{WEIGHT_COMPONENTS[reference][1]}</div>
            <div style="font-size: 0.875rem; color: #92400e;">Raw: 100 → Normalized: {normalized[reference]:.4f}</div>
        </div>
        """, unsafe_allow_html=True)
        ordered = sorted([(k, v) for k, v in normalized.items() if k != reference], key=lambda x: x[1], reverse=True)
        for key, wgt in ordered:
            name = WEIGHT_COMPONENTS[key][1]
            locked = key in inactive_keys
            color = "#9ca3af" if locked else ("#667eea" if wgt >= 0.10 else "#f59e0b" if wgt >= 0.05 else "#6b7280")
            sub = "locked (inactive)" if locked else f"raw: {raw[key]:.0f}"
            st.markdown(f"""
            <div style="background: white; padding: 0.75rem; border-radius: 8px; border-left: 4px solid {color};
                        margin-bottom: 0.5rem; box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div><div style="font-weight: 600; color: #1f2937;">{name}</div>
                         <div style="font-size: 0.875rem; color: #6b7280;">{WEIGHT_NOTATION[key]} ({sub})</div></div>
                    <div style="text-align: right;">
                        <div style="font-size: 1.5rem; font-weight: 700; color: {color};">{wgt:.4f}</div>
                        <div style="font-size: 0.75rem; color: #10b981;">{wgt*100:.1f}%</div></div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        st.markdown(f"""
        <div style="background: #f3f4f6; padding: 1rem; border-radius: 8px; margin-top: 1rem;">
            <strong>Total Sum:</strong> {sum(normalized.values()):.10f}
        </div>
        """, unsafe_allow_html=True)

    st.success("✅ Weights configured using SWING weighting. Click 'Next' to run optimization.")


# ================================================================
# STEP 4: RUN OPTIMIZATION
# ================================================================

def show_step4_run_optimization():
    st.header("🚀 Step 4: Run Optimization")
    data, weights = st.session_state.data, st.session_state.weights
    if not data or not weights:
        st.warning("⚠️ Please complete previous steps first!")
        return

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("🚀 Run Optimization", type="primary", use_container_width=True):
            with st.spinner("Building and solving the criterion-selection MILP..."):
                try:
                    solution, gates, reward_coeff, denominators, method = solve_model(data, weights)
                    frames = build_result_frames(data, weights, solution, gates, reward_coeff, denominators, method)
                    st.session_state.solution = solution
                    st.session_state.result_frames = frames
                except Exception as e:
                    st.session_state.solution = None
                    st.session_state.result_frames = None
                    st.error(f"❌ {str(e)}")
                    st.info("If the model is infeasible, consider relaxing thresholds, representation targets, or the Assessment Mode interval, or deactivating a property in Step 1.")
                    return

    solution, frames = st.session_state.solution, st.session_state.result_frames
    if not solution:
        return

    st.markdown('<div class="success-box"><strong>✅ Optimization completed successfully!</strong></div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    c1.metric("Selected Criteria (N)", f"{solution['N']}/{data['num_criteria']}")
    c2.metric("Quantitative Proportion (rho)", f"{solution['rho']:.4f}")
    c3.metric("Objective Value (Z)", f"{solution['objective']:.6f}")
    st.caption(frames['Summary'].loc[1, 'Value'])

    st.subheader("✅ Selected Criteria")
    st.dataframe(frames['Selected_Portfolio'], use_container_width=True, hide_index=True)

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    d1, d2 = st.columns(2)
    with d1:
        st.download_button("📥 Download Full Results (Excel, 8 sheets)", data=export_results_excel(frames),
                           file_name=f"CREST_Optimization_Results_{stamp}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, type="primary")
    with d2:
        st.download_button("📥 Download Selected Criteria (CSV)", data=frames['Selected_Portfolio'].to_csv(index=False),
                           file_name=f"selected_criteria_{stamp}.csv", mime="text/csv", use_container_width=True)

    with st.expander("📊 Criterion decisions (gates, conflicts, and selection status)", expanded=True):
        cols = ["ID", "Criterion", "x_i", "Final status", "Primary exclusion reason", "Represented objectives", "Assessment Mode", "Normalized reward contribution"]
        st.dataframe(frames['Criterion_Decisions'][cols], use_container_width=True, hide_index=True)
    with st.expander("🎯 Objective coverage"):
        st.dataframe(frames['Objective_Coverage'], use_container_width=True, hide_index=True)
    with st.expander("🧮 Objective function breakdown"):
        st.dataframe(frames['Objective_Breakdown'], use_container_width=True, hide_index=True)
        st.write(f"**Reward:** {solution['reward']:.6f}  |  **Representativeness penalty:** {solution['rep_penalty']:.6f}  |  "
                 f"**Parsimony penalty:** {solution['parsimony_penalty']:.6f}  |  **Distinctiveness penalty:** {solution['distinctiveness_penalty']:.6f}  |  "
                 f"**Z:** {solution['objective']:.6f}")
    if data['active'][12]:
        with st.expander("🔗 Pairwise review (Distinctiveness)"):
            st.dataframe(frames['Pairwise_Review'], use_container_width=True, hide_index=True)
    with st.expander("⚙️ Model controls used"):
        st.dataframe(frames['Model_Controls'], use_container_width=True, hide_index=True)


# ================================================================
# MAIN APPLICATION
# ================================================================

def main():
    st.markdown('<h1 class="main-title">Criteria Retrieval, Extraction and Selection Tool (CREST)</h1>', unsafe_allow_html=True)
    st.markdown('<p class="sub-title">Phase 2 Property Examination and Phase 3 Criteria Selection Optimization (13 properties)</p>', unsafe_allow_html=True)

    show_progress_indicator(st.session_state.current_step)
    st.markdown("---")

    with st.sidebar:
        st.markdown("### 📊 Problem Information")
        data = st.session_state.data
        if data:
            c1, c2 = st.columns(2)
            with c1:
                st.metric("Criteria", data['num_criteria'])
                st.metric("Experts", data['num_experts'])
            with c2:
                st.metric("Alternatives", data['num_alternatives'])
                st.metric("Objectives", data['num_objectives'])
            n_active = sum(1 for p in range(1, 14) if data['active'][p])
            st.markdown(f"**Active properties:** {n_active}/13")
        else:
            st.info("Upload data to see problem details")
        st.markdown("---")
        st.markdown("### 🧭 Quick Navigation")
        if st.button("📝 Step 1: Generate", use_container_width=True, type="secondary"):
            st.session_state.current_step = 1; st.rerun()
        if st.button("📤 Step 2: Upload", use_container_width=True, type="secondary"):
            st.session_state.current_step = 2; st.rerun()
        if st.button("⚖️ Step 3: Weights", use_container_width=True, type="secondary", disabled=not st.session_state.data):
            st.session_state.current_step = 3; st.rerun()
        if st.button("🚀 Step 4: Optimize", use_container_width=True, type="secondary", disabled=not st.session_state.weights):
            st.session_state.current_step = 4; st.rerun()

    step = st.session_state.current_step
    if step == 1:
        show_step1_generate_template()
    elif step == 2:
        show_step2_upload_extract()
    elif step == 3:
        show_step3_set_weights()
    elif step == 4:
        show_step4_run_optimization()

    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown("---")
    show_navigation_buttons(step)


if __name__ == "__main__":
    main()
